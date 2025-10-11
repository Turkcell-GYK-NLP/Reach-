#!/usr/bin/env python3
"""
Tarife Önerisi Sistemi - Pandas Olmadan
Kullanıcı verilerine göre en uygun tarife önerisi yapan sistem
"""

from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional
import json
import random

class TarifeOnerisiSistemi:
    def __init__(self):
        """Tarife önerisi sistemi başlatıcı"""
        self.tarifeler = self._load_tarifeler()
        self.kullanici_verileri = None
        
    def _load_tarifeler(self) -> List[Dict]:
        """Güncel tarife bilgilerini yükler"""
        # Önce kapsamlı JSON dosyasından yüklemeyi dene
        try:
            with open('guncel_tarifeler_2025_kapsamli.json', 'r', encoding='utf-8') as f:
                tarifeler_data = json.load(f)
                all_tarifeler = []
                for operator, plans in tarifeler_data.items():
                    all_tarifeler.extend(plans)
                return all_tarifeler
        except FileNotFoundError:
            pass

        # Eski JSON dosyasından yüklemeyi dene
        try:
            with open('guncel_tarifeler_2025.json', 'r', encoding='utf-8') as f:
                tarifeler_data = json.load(f)
                all_tarifeler = []
                for operator, plans in tarifeler_data.items():
                    all_tarifeler.extend(plans)
                return all_tarifeler
        except FileNotFoundError:
            pass

        # JSON dosyası yoksa varsayılan verileri kullan
        tarifeler = {
            'Turkcell': [
                {'ad': 'Turkcell 5GB', 'data_gb': 5, 'dakika': 200, 'sms': 100, 'fiyat': 185.0, 'operator': 'Turkcell'},
                {'ad': 'Turkcell 10GB', 'data_gb': 10, 'dakika': 500, 'sms': 200, 'fiyat': 250.0, 'operator': 'Turkcell'},
                {'ad': 'Turkcell 20GB', 'data_gb': 20, 'dakika': 1000, 'sms': 500, 'fiyat': 350.0, 'operator': 'Turkcell'},
                {'ad': 'Turkcell 50GB', 'data_gb': 50, 'dakika': 2000, 'sms': 1000, 'fiyat': 500.0, 'operator': 'Turkcell'},
                {'ad': 'Turkcell Sınırsız', 'data_gb': 999, 'dakika': 2000, 'sms': 1000, 'fiyat': 700.0, 'operator': 'Turkcell'}
            ],
            'Vodafone': [
                {'ad': 'Vodafone 5GB', 'data_gb': 5, 'dakika': 200, 'sms': 100, 'fiyat': 175.0, 'operator': 'Vodafone'},
                {'ad': 'Vodafone 10GB', 'data_gb': 10, 'dakika': 500, 'sms': 200, 'fiyat': 230.0, 'operator': 'Vodafone'},
                {'ad': 'Vodafone 20GB', 'data_gb': 20, 'dakika': 1000, 'sms': 500, 'fiyat': 300.0, 'operator': 'Vodafone'},
                {'ad': 'Vodafone 50GB', 'data_gb': 50, 'dakika': 2000, 'sms': 1000, 'fiyat': 420.0, 'operator': 'Vodafone'},
                {'ad': 'Vodafone Sınırsız', 'data_gb': 999, 'dakika': 2000, 'sms': 1000, 'fiyat': 600.0, 'operator': 'Vodafone'}
            ],
            'Türk Telekom': [
                {'ad': 'Türk Telekom 5GB', 'data_gb': 5, 'dakika': 200, 'sms': 100, 'fiyat': 180.0, 'operator': 'Türk Telekom'},
                {'ad': 'Türk Telekom 10GB', 'data_gb': 10, 'dakika': 500, 'sms': 200, 'fiyat': 240.0, 'operator': 'Türk Telekom'},
                {'ad': 'Türk Telekom 20GB', 'data_gb': 20, 'dakika': 1000, 'sms': 500, 'fiyat': 320.0, 'operator': 'Türk Telekom'},
                {'ad': 'Türk Telekom 50GB', 'data_gb': 50, 'dakika': 2000, 'sms': 1000, 'fiyat': 450.0, 'operator': 'Türk Telekom'},
                {'ad': 'Türk Telekom Sınırsız', 'data_gb': 999, 'dakika': 2000, 'sms': 1000, 'fiyat': 650.0, 'operator': 'Türk Telekom'}
            ]
        }
        
        all_tarifeler = []
        for operator, plans in tarifeler.items():
            all_tarifeler.extend(plans)
        
        return all_tarifeler
    
    def load_kullanici_verileri(self, excel_path: str):
        """Excel dosyasından kullanıcı verilerini yükler"""
        try:
            # Excel dosyasını openpyxl ile oku
            workbook = load_workbook(excel_path)
            worksheet = workbook.active
            
            # Başlık satırını oku
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value)
            
            # Veri satırlarını oku
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Boş satırları atla
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = value
                    data.append(row_dict)
            
            self.kullanici_verileri = data
            print(f"✅ {len(self.kullanici_verileri)} kullanıcı verisi yüklendi")
            return True
        except Exception as e:
            print(f"❌ Veri yükleme hatası: {e}")
            return False
    
    def _calculate_tarife_uygunlugu(self, kullanici: Dict, tarife: Dict) -> Tuple[bool, float, str]:
        """
        Kullanıcı için tarife uygunluğunu hesaplar
        Returns: (uygun_mu, skor, aciklama)
        """
        # Kullanıcı ihtiyaçları
        ihtiyac_data = kullanici.get('monthly_data_gb', 0)
        ihtiyac_dakika = kullanici.get('monthly_calls_min', 0)
        ihtiyac_sms = kullanici.get('monthly_sms', 0)

        # Tarife içerikleri
        tarife_data = tarife['data_gb']
        tarife_dakika = tarife['dakika']
        tarife_sms = tarife['sms']

        # Uygunluk kontrolü
        data_uygun = tarife_data >= ihtiyac_data
        dakika_uygun = tarife_dakika >= ihtiyac_dakika
        sms_uygun = tarife_sms >= ihtiyac_sms

        if not (data_uygun and dakika_uygun and sms_uygun):
            eksikler = []
            if not data_uygun:
                eksikler.append(f"Veri ({ihtiyac_data}GB ihtiyaç, {tarife_data}GB mevcut)")
            if not dakika_uygun:
                eksikler.append(f"Dakika ({ihtiyac_dakika}dk ihtiyaç, {tarife_dakika}dk mevcut)")
            if not sms_uygun:
                eksikler.append(f"SMS ({ihtiyac_sms} ihtiyaç, {tarife_sms} mevcut)")

            return False, 0, f"Uygun değil: {', '.join(eksikler)}"

        # Kapsamlı skor hesaplama
        # 1. Fiyat skoru (30% ağırlık)
        fiyat_skoru = 1000 / tarife['fiyat']  # Düşük fiyat = yüksek skor

        # 2. Kota verimliliği (20% ağırlık)
        data_verim = 1 - abs(tarife_data - ihtiyac_data) / max(tarife_data, ihtiyac_data)
        dakika_verim = 1 - abs(tarife_dakika - ihtiyac_dakika) / max(tarife_dakika, ihtiyac_dakika)
        sms_verim = 1 - abs(tarife_sms - ihtiyac_sms) / max(tarife_sms, ihtiyac_sms)
        verim_skoru = (data_verim + dakika_verim + sms_verim) / 3

        # 3. Operatör kalite skoru (50% ağırlık)
        kalite_skoru = 0
        if 'genel_puan' in tarife:
            kalite_skoru = tarife['genel_puan']
        else:
            # Eski veriler için varsayılan kalite skoru
            operator = tarife['operator']
            if operator == 'Turkcell':
                kalite_skoru = 8.5
            elif operator == 'Türk Telekom':
                kalite_skoru = 8.0
            elif operator == 'Vodafone':
                kalite_skoru = 7.5

        # Ağırlıklı toplam skor
        toplam_skor = (fiyat_skoru * 0.3 + verim_skoru * 0.2 + kalite_skoru * 0.5)

        # Açıklama oluştur
        kalite_detay = ""
        if 'cekim_guclu' in tarife:
            kalite_detay = f", Çekim: {tarife['cekim_guclu']}/10, Hız: {tarife['hiz']}/10, Güvenilirlik: {tarife['guvenilirlik']}/10"

        aciklama = f"Uygun - Fiyat: {tarife['fiyat']}TL, Verim: {verim_skoru:.2f}, Kalite: {kalite_skoru:.1f}/10{kalite_detay}"

        return True, toplam_skor, aciklama

    def _find_best_tarife(self, kullanici: Dict) -> Tuple[Optional[Dict], float, str]:
        """Kullanıcı için en uygun tarifeyi bulur"""
        uygun_tarifeler = []
        
        for tarife in self.tarifeler:
            uygun, skor, aciklama = self._calculate_tarife_uygunlugu(kullanici, tarife)
            if uygun:
                uygun_tarifeler.append((tarife, skor, aciklama))
        
        if not uygun_tarifeler:
            return None, 0, "Uygun tarife bulunamadı"
        
        # En yüksek skorlu tarifeyi seç
        en_iyi = max(uygun_tarifeler, key=lambda x: x[1])
        return en_iyi[0], en_iyi[1], en_iyi[2]

    def get_tarife_onerisi(self, user_id: int) -> Dict:
        """Belirli bir kullanıcı için tarife önerisi"""
        if self.kullanici_verileri is None:
            return {"hata": "Kullanıcı verileri yüklenmemiş"}
        
        # Kullanıcıyı bul
        kullanici = None
        for user in self.kullanici_verileri:
            if user.get('user_id') == user_id:
                kullanici = user
                break
        
        if not kullanici:
            return {"hata": "Kullanıcı bulunamadı"}
        
        # En uygun tarifeyi bul
        en_uygun_tarife, skor, aciklama = self._find_best_tarife(kullanici)
        
        if not en_uygun_tarife:
            return {"hata": "Uygun tarife bulunamadı"}
        
        # Alternatif tarifeleri bul
        alternatif_tarifeler = []
        for tarife in self.tarifeler:
            uygun, skor, aciklama = self._calculate_tarife_uygunlugu(kullanici, tarife)
            if uygun and tarife != en_uygun_tarife:
                alternatif_tarifeler.append({
                    'tarife': tarife,
                    'skor': skor,
                    'aciklama': aciklama
                })
        
        # Skora göre sırala
        alternatif_tarifeler.sort(key=lambda x: x['skor'], reverse=True)
        
        return {
            'kullanici_id': user_id,
            'en_uygun_tarife': en_uygun_tarife,
            'skor': skor,
            'aciklama': aciklama,
            'kullanici_ihtiyaclari': {
                'monthly_data_gb': kullanici.get('monthly_data_gb', 0),
                'monthly_calls_min': kullanici.get('monthly_calls_min', 0),
                'monthly_sms': kullanici.get('monthly_sms', 0)
            },
            'alternatif_tarifeler': alternatif_tarifeler[:5]  # En iyi 5 alternatif
        }

    def profil_bazli_analiz(self) -> Dict:
        """Kullanım profillerine göre analiz yapar"""
        if self.kullanici_verileri is None:
            return {"hata": "Kullanıcı verileri yüklenmemiş"}
        
        # Kullanım profillerini oluştur
        profiles = {'Light User': [], 'Balanced': [], 'Data Heavy': []}
        
        for user in self.kullanici_verileri:
            data_gb = user.get('monthly_data_gb', 0)
            if data_gb >= 30:
                profiles['Data Heavy'].append(user)
            elif data_gb >= 15:
                profiles['Data Heavy'].append(user)
            elif data_gb >= 5:
                profiles['Balanced'].append(user)
            else:
                profiles['Light User'].append(user)
        
        # Profil bazlı analiz
        profile_analysis = {}
        for profile_name, users in profiles.items():
            if not users:
                continue
                
            # Rastgele bir kullanıcı seç
            random_user = random.choice(users)
            
            # Tarife önerisi hesapla
            en_uygun_tarife, skor, aciklama = self._find_best_tarife(random_user)
            
            profile_analysis[profile_name] = {
                'kullanici_sayisi': len(users),
                'ornek_kullanici': random_user,
                'en_uygun_tarife': en_uygun_tarife,
                'skor': skor,
                'aciklama': aciklama
            }
        
        return profile_analysis

def main():
    import sys
    
    # Sistem başlat
    sistem = TarifeOnerisiSistemi()
    
    # Excel dosyasını yükle
    excel_path = 'usage_with_recommendations.xlsx'
    if not sistem.load_kullanici_verileri(excel_path):
        print("Excel dosyası yüklenemedi, varsayılan verilerle devam ediliyor...")
    
    # Komut satırı argümanlarını kontrol et
    if len(sys.argv) > 1:
        if sys.argv[1] == '--user-id' and len(sys.argv) > 2:
            user_id = int(sys.argv[2])
            sonuc = sistem.get_tarife_onerisi(user_id)
            print(json.dumps(sonuc, ensure_ascii=False, indent=2))
        elif sys.argv[1] == '--bulk-analysis':
            sample_size = int(sys.argv[2]) if len(sys.argv) > 2 else 100
            sonuc = sistem.toplu_analiz(sample_size)
            print(json.dumps(sonuc, ensure_ascii=False, indent=2))
        elif sys.argv[1] == '--profile-analysis':
            sonuc = sistem.profil_bazli_analiz()
            print(json.dumps(sonuc, ensure_ascii=False, indent=2))
        else:
            print("Geçersiz argüman")
    else:
        print("Kullanım: python tarife_onerisi_sistemi.py [--user-id ID | --bulk-analysis [SIZE] | --profile-analysis]")

if __name__ == "__main__":
    main()
