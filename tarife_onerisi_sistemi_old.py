#!/usr/bin/env python3
"""
Tarife Önerisi Sistemi
Kullanıcı verilerine göre en uygun tarife önerisi yapan sistem
"""

from openpyxl import load_workbook
from typing import Dict, List, Tuple, Optional
import json
import random

class TarifeOnerisiSistemi:
    def __init__(self):
        """Tarife önerisi sistemi başlatıcı"""
        self.tarifeler = self._load_tarifeler()
        self.kullanici_verileri = None
        
    def _load_tarifeler(self) -> List[Dict]:
        """Güncel tarife bilgilerini yükler"""
        # Önce kapsamlı JSON dosyasından yüklemeyi dene
        try:
            with open('guncel_tarifeler_2025_kapsamli.json', 'r', encoding='utf-8') as f:
                tarifeler_data = json.load(f)
                all_tarifeler = []
                for operator, plans in tarifeler_data.items():
                    all_tarifeler.extend(plans)
                return all_tarifeler
        except FileNotFoundError:
            pass
        
        # Eski JSON dosyasından yüklemeyi dene
        try:
            with open('guncel_tarifeler_2025.json', 'r', encoding='utf-8') as f:
                tarifeler_data = json.load(f)
                all_tarifeler = []
                for operator, plans in tarifeler_data.items():
                    all_tarifeler.extend(plans)
                return all_tarifeler
        except FileNotFoundError:
            pass
        
        # JSON dosyası yoksa varsayılan verileri kullan
        tarifeler = {
            'Turkcell': [
                {'ad': 'Turkcell 5GB', 'data_gb': 5, 'dakika': 200, 'sms': 100, 'fiyat': 185.0, 'operator': 'Turkcell'},
                {'ad': 'Turkcell 10GB', 'data_gb': 10, 'dakika': 500, 'sms': 200, 'fiyat': 250.0, 'operator': 'Turkcell'},
                {'ad': 'Turkcell 20GB', 'data_gb': 20, 'dakika': 1000, 'sms': 500, 'fiyat': 350.0, 'operator': 'Turkcell'},
                {'ad': 'Turkcell 50GB', 'data_gb': 50, 'dakika': 2000, 'sms': 1000, 'fiyat': 500.0, 'operator': 'Turkcell'},
                {'ad': 'Turkcell Sınırsız', 'data_gb': 999, 'dakika': 2000, 'sms': 1000, 'fiyat': 700.0, 'operator': 'Turkcell'}
            ],
            'Vodafone': [
                {'ad': 'Vodafone 5GB', 'data_gb': 5, 'dakika': 200, 'sms': 100, 'fiyat': 175.0, 'operator': 'Vodafone'},
                {'ad': 'Vodafone 10GB', 'data_gb': 10, 'dakika': 500, 'sms': 200, 'fiyat': 230.0, 'operator': 'Vodafone'},
                {'ad': 'Vodafone 20GB', 'data_gb': 20, 'dakika': 1000, 'sms': 500, 'fiyat': 300.0, 'operator': 'Vodafone'},
                {'ad': 'Vodafone 50GB', 'data_gb': 50, 'dakika': 2000, 'sms': 1000, 'fiyat': 420.0, 'operator': 'Vodafone'},
                {'ad': 'Vodafone Sınırsız', 'data_gb': 999, 'dakika': 2000, 'sms': 1000, 'fiyat': 600.0, 'operator': 'Vodafone'}
            ],
            'Türk Telekom': [
                {'ad': 'Türk Telekom 5GB', 'data_gb': 5, 'dakika': 200, 'sms': 100, 'fiyat': 180.0, 'operator': 'Türk Telekom'},
                {'ad': 'Türk Telekom 10GB', 'data_gb': 10, 'dakika': 500, 'sms': 200, 'fiyat': 240.0, 'operator': 'Türk Telekom'},
                {'ad': 'Türk Telekom 20GB', 'data_gb': 20, 'dakika': 1000, 'sms': 500, 'fiyat': 320.0, 'operator': 'Türk Telekom'},
                {'ad': 'Türk Telekom 50GB', 'data_gb': 50, 'dakika': 2000, 'sms': 1000, 'fiyat': 450.0, 'operator': 'Türk Telekom'},
                {'ad': 'Türk Telekom Sınırsız', 'data_gb': 999, 'dakika': 2000, 'sms': 1000, 'fiyat': 650.0, 'operator': 'Türk Telekom'}
            ]
        }
        
        # Tüm tarifeleri tek listede topla
        all_tarifeler = []
        for operator, plans in tarifeler.items():
            all_tarifeler.extend(plans)
        
        return all_tarifeler
    
    def load_kullanici_verileri(self, excel_path: str):
        """Excel dosyasından kullanıcı verilerini yükler"""
        try:
            # Excel dosyasını openpyxl ile oku
            workbook = load_workbook(excel_path)
            worksheet = workbook.active
            
            # Başlık satırını oku
            headers = []
            for cell in worksheet[1]:
                headers.append(cell.value)
            
            # Veri satırlarını oku
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Boş satırları atla
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            row_dict[headers[i]] = value
                    data.append(row_dict)
            
            self.kullanici_verileri = data
            print(f"✅ {len(self.kullanici_verileri)} kullanıcı verisi yüklendi")
            return True
        except Exception as e:
            print(f"❌ Veri yükleme hatası: {e}")
            return False
    
    def _calculate_tarife_uygunlugu(self, kullanici: Dict, tarife: Dict) -> Tuple[bool, float, str]:
        """
        Kullanıcı için tarife uygunluğunu hesaplar
        Returns: (uygun_mu, skor, aciklama)
        """
        # Kullanıcı ihtiyaçları
        ihtiyac_data = kullanici['monthly_data_gb']
        ihtiyac_dakika = kullanici['monthly_calls_min']
        ihtiyac_sms = kullanici['monthly_sms']
        
        # Tarife içerikleri
        tarife_data = tarife['data_gb']
        tarife_dakika = tarife['dakika']
        tarife_sms = tarife['sms']
        
        # Uygunluk kontrolü
        data_uygun = tarife_data >= ihtiyac_data
        dakika_uygun = tarife_dakika >= ihtiyac_dakika
        sms_uygun = tarife_sms >= ihtiyac_sms
        
        if not (data_uygun and dakika_uygun and sms_uygun):
            eksikler = []
            if not data_uygun:
                eksikler.append(f"Veri ({ihtiyac_data}GB ihtiyaç, {tarife_data}GB mevcut)")
            if not dakika_uygun:
                eksikler.append(f"Dakika ({ihtiyac_dakika}dk ihtiyaç, {tarife_dakika}dk mevcut)")
            if not sms_uygun:
                eksikler.append(f"SMS ({ihtiyac_sms} ihtiyaç, {tarife_sms} mevcut)")
            
            return False, 0, f"Uygun değil: {', '.join(eksikler)}"
        
        # Kapsamlı skor hesaplama
        # 1. Fiyat skoru (30% ağırlık)
        fiyat_skoru = 1000 / tarife['fiyat']  # Düşük fiyat = yüksek skor
        
        # 2. Kota verimliliği (20% ağırlık)
        data_verim = 1 - abs(tarife_data - ihtiyac_data) / max(tarife_data, ihtiyac_data)
        dakika_verim = 1 - abs(tarife_dakika - ihtiyac_dakika) / max(tarife_dakika, ihtiyac_dakika)
        sms_verim = 1 - abs(tarife_sms - ihtiyac_sms) / max(tarife_sms, ihtiyac_sms)
        verim_skoru = (data_verim + dakika_verim + sms_verim) / 3
        
        # 3. Operatör kalite skoru (50% ağırlık)
        kalite_skoru = 0
        if 'genel_puan' in tarife:
            kalite_skoru = tarife['genel_puan']
        else:
            # Eski veriler için varsayılan kalite skoru
            operator = tarife['operator']
            if operator == 'Turkcell':
                kalite_skoru = 8.5
            elif operator == 'Türk Telekom':
                kalite_skoru = 8.0
            elif operator == 'Vodafone':
                kalite_skoru = 7.5
        
        # Ağırlıklı toplam skor
        toplam_skor = (fiyat_skoru * 0.3 + verim_skoru * 0.2 + kalite_skoru * 0.5)
        
        # Açıklama oluştur
        kalite_detay = ""
        if 'cekim_guclu' in tarife:
            kalite_detay = f", Çekim: {tarife['cekim_guclu']}/10, Hız: {tarife['hiz']}/10, Güvenilirlik: {tarife['guvenilirlik']}/10"
        
        aciklama = f"Uygun - Fiyat: {tarife['fiyat']}TL, Verim: {verim_skoru:.2f}, Kalite: {kalite_skoru:.1f}/10{kalite_detay}"
        
        return True, toplam_skor, aciklama
    
    def kullanici_icin_en_uygun_tarife(self, kullanici_id: int) -> Dict:
        """Belirli bir kullanıcı için en uygun tarifeyi bulur"""
        if self.kullanici_verileri is None:
            return {"hata": "Kullanıcı verileri yüklenmemiş"}
        
        kullanici = self.kullanici_verileri[self.kullanici_verileri['user_id'] == kullanici_id]
        if kullanici.empty:
            return {"hata": f"Kullanıcı {kullanici_id} bulunamadı"}
        
        kullanici = kullanici.iloc[0].to_dict()
        
        # Tüm tarifeleri değerlendir
        uygun_tarifeler = []
        
        for tarife in self.tarifeler:
            uygun, skor, aciklama = self._calculate_tarife_uygunlugu(kullanici, tarife)
            if uygun:
                uygun_tarifeler.append({
                    'tarife': tarife,
                    'skor': skor,
                    'aciklama': aciklama
                })
        
        if not uygun_tarifeler:
            return {
                "kullanici_id": kullanici_id,
                "hata": "Hiçbir tarife uygun değil",
                "kullanici_ihtiyaclari": {
                    "data_gb": kullanici['monthly_data_gb'],
                    "dakika": kullanici['monthly_calls_min'],
                    "sms": kullanici['monthly_sms']
                }
            }
        
        # En yüksek skorlu tarifeyi seç
        en_iyi = max(uygun_tarifeler, key=lambda x: x['skor'])
        
        return {
            "kullanici_id": kullanici_id,
            "en_uygun_tarife": en_iyi['tarife'],
            "skor": en_iyi['skor'],
            "aciklama": en_iyi['aciklama'],
            "kullanici_ihtiyaclari": {
                "data_gb": kullanici['monthly_data_gb'],
                "dakika": kullanici['monthly_calls_min'],
                "sms": kullanici['monthly_sms']
            },
            "alternatif_tarifeler": sorted(uygun_tarifeler, key=lambda x: x['skor'], reverse=True)[1:4]  # İlk 3 alternatif
        }
    
    def toplu_analiz(self, sample_size: int = 100) -> Dict:
        """Tüm kullanıcılar için toplu analiz yapar"""
        if self.kullanici_verileri is None:
            return {"hata": "Kullanıcı verileri yüklenmemiş"}
        
        # Örnekleme yap
        sample_users = self.kullanici_verileri.sample(n=min(sample_size, len(self.kullanici_verileri)))
        
        sonuclar = []
        operator_tercihleri = {'Turkcell': 0, 'Vodafone': 0, 'Türk Telekom': 0}
        fiyat_tasarrufu = 0
        
        for _, kullanici in sample_users.iterrows():
            sonuc = self.kullanici_icin_en_uygun_tarife(kullanici['user_id'])
            if 'hata' not in sonuc:
                sonuclar.append(sonuc)
                operator_tercihleri[sonuc['en_uygun_tarife']['operator']] += 1
                
                # Fiyat tasarrufu hesapla
                mevcut_fatura = kullanici['avg_bill_placeholder_tl']
                onerilen_fiyat = sonuc['en_uygun_tarife']['fiyat']
                fiyat_tasarrufu += max(0, mevcut_fatura - onerilen_fiyat)
        
        return {
            "analiz_edilen_kullanici_sayisi": len(sonuclar),
            "operator_tercihleri": operator_tercihleri,
            "ortalama_fiyat_tasarrufu": fiyat_tasarrufu / len(sonuclar) if sonuclar else 0,
            "toplam_fiyat_tasarrufu": fiyat_tasarrufu,
            "ornek_sonuclar": sonuclar[:10]  # İlk 10 sonuç
        }
    
    def profil_bazli_analiz(self) -> Dict:
        """Kullanım profillerine göre analiz yapar"""
        if self.kullanici_verileri is None:
            return {"hata": "Kullanıcı verileri yüklenmemiş"}
        
        # Kullanım profillerini oluştur
        df = self.kullanici_verileri.copy()
        df['data_category'] = pd.cut(df['monthly_data_gb'], 
                                    bins=[0, 5, 15, 30, float('inf')], 
                                    labels=['Düşük (0-5GB)', 'Orta (5-15GB)', 'Yüksek (15-30GB)', 'Çok Yüksek (30GB+)'])
        
        def create_usage_profile(row):
            data_cat = row['data_category']
            if data_cat in ['Çok Yüksek (30GB+)']:
                return 'Data Heavy'
            elif data_cat in ['Yüksek (15-30GB)']:
                return 'Data Heavy'
            elif data_cat in ['Orta (5-15GB)']:
                return 'Balanced'
            else:
                return 'Light User'
        
        df['usage_profile'] = df.apply(create_usage_profile, axis=1)
        
        profil_analizleri = {}
        
        for profile in df['usage_profile'].unique():
            profile_users = df[df['usage_profile'] == profile]
            sample_size = min(50, len(profile_users))  # Her profilden max 50 kullanıcı
            sample_users = profile_users.sample(n=sample_size)
            
            profil_sonuclari = []
            operator_tercihleri = {'Turkcell': 0, 'Vodafone': 0, 'Türk Telekom': 0}
            
            for _, kullanici in sample_users.iterrows():
                sonuc = self.kullanici_icin_en_uygun_tarife(kullanici['user_id'])
                if 'hata' not in sonuc:
                    profil_sonuclari.append(sonuc)
                    operator_tercihleri[sonuc['en_uygun_tarife']['operator']] += 1
            
            profil_analizleri[profile] = {
                "kullanici_sayisi": len(profil_sonuclari),
                "operator_tercihleri": operator_tercihleri,
                "ortalama_ihtiyaclar": {
                    "data_gb": profile_users['monthly_data_gb'].mean(),
                    "dakika": profile_users['monthly_calls_min'].mean(),
                    "sms": profile_users['monthly_sms'].mean()
                },
                "ornek_sonuclar": profil_sonuclari[:5]
            }
        
        return profil_analizleri

def main():
    """Ana fonksiyon - komut satırı argümanlarını işler"""
    import sys
    import json
    
    # Sistemi başlat
    sistem = TarifeOnerisiSistemi()
    
    # Kullanıcı verilerini yükle
    if not sistem.load_kullanici_verileri('usage_with_recommendations.xlsx'):
        return
    
    # Komut satırı argümanlarını kontrol et
    if len(sys.argv) > 1:
        if '--user-id' in sys.argv:
            user_id_index = sys.argv.index('--user-id') + 1
            if user_id_index < len(sys.argv):
                user_id = int(sys.argv[user_id_index])
                sonuc = sistem.kullanici_icin_en_uygun_tarife(user_id)
                print(json.dumps(sonuc, ensure_ascii=False, indent=2))
                return
        
        elif '--bulk-analysis' in sys.argv:
            sample_size_index = sys.argv.index('--bulk-analysis') + 1
            sample_size = 100
            if sample_size_index < len(sys.argv):
                sample_size = int(sys.argv[sample_size_index])
            
            toplu_sonuc = sistem.toplu_analiz(sample_size=sample_size)
            print(json.dumps(toplu_sonuc, ensure_ascii=False, indent=2))
            return
        
        elif '--profile-analysis' in sys.argv:
            profil_analizleri = sistem.profil_bazli_analiz()
            print(json.dumps(profil_analizleri, ensure_ascii=False, indent=2))
            return
    
    # Varsayılan analiz (komut satırı argümanı yoksa)
    print("=== TARİFE ÖNERİSİ SİSTEMİ ===\n")
    
    print("\n=== ÖRNEK KULLANICI ANALİZİ ===")
    # İlk 5 kullanıcı için analiz
    for user_id in [1, 2, 3, 4, 5]:
        sonuc = sistem.kullanici_icin_en_uygun_tarife(user_id)
        print(f"\nKullanıcı {user_id}:")
        if 'hata' not in sonuc:
            tarife = sonuc['en_uygun_tarife']
            ihtiyac = sonuc['kullanici_ihtiyaclari']
            print(f"  İhtiyaçlar: {ihtiyac['data_gb']:.1f}GB, {ihtiyac['dakika']}dk, {ihtiyac['sms']}SMS")
            print(f"  Önerilen: {tarife['ad']} - {tarife['fiyat']}TL")
            print(f"  Açıklama: {sonuc['aciklama']}")
        else:
            print(f"  Hata: {sonuc['hata']}")
    
    print("\n=== TOPLU ANALİZ ===")
    toplu_sonuc = sistem.toplu_analiz(sample_size=100)
    print(f"Analiz edilen kullanıcı sayısı: {toplu_sonuc['analiz_edilen_kullanici_sayisi']}")
    print(f"Ortalama fiyat tasarrufu: {toplu_sonuc['ortalama_fiyat_tasarrufu']:.2f} TL")
    print("Operatör tercihleri:")
    for operator, count in toplu_sonuc['operator_tercihleri'].items():
        print(f"  {operator}: {count} kullanıcı")
    
    print("\n=== PROFİL BAZLI ANALİZ ===")
    profil_analizleri = sistem.profil_bazli_analiz()
    for profile, analiz in profil_analizleri.items():
        print(f"\n{profile} Profili:")
        print(f"  Kullanıcı sayısı: {analiz['kullanici_sayisi']}")
        print(f"  Ortalama ihtiyaçlar: {analiz['ortalama_ihtiyaclar']['data_gb']:.1f}GB, {analiz['ortalama_ihtiyaclar']['dakika']}dk, {analiz['ortalama_ihtiyaclar']['sms']}SMS")
        print("  Operatör tercihleri:")
        for operator, count in analiz['operator_tercihleri'].items():
            print(f"    {operator}: {count} kullanıcı")

if __name__ == "__main__":
    main()
